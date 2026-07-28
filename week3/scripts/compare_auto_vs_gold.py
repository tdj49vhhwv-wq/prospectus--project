#!/usr/bin/env python3
"""
Auto vs Gold 逐字段对比

1. 运行 auto_extract 得到自动提取结果
2. 加载 Gold JSONL
3. 按 (日期, 名称) 对齐记录
4. 逐字段 diff，标记 match / mismatch / 漏提 / 误提 / 待复核

输出: reports/auto_vs_gold_comparison.xlsx
"""
import sys
import json
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from auto_extract import auto_extract_all, COMPANY_MD_MAP
    HAS_AUTO = True
except ImportError:
    HAS_AUTO = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XL = True
except ImportError:
    HAS_XL = False

GOLD_DIR = PROJECT_ROOT / "outputs" / "week2_jsonl"
REPORTS_DIR = PROJECT_ROOT / "reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

# Field lists for comparison
SF_FIELDS = [
    "subscription_date", "subscriber_name", "shares_subscribed",
    "amount_subscribed", "price_per_share", "event_context",
    "post_event_total_shares", "post_event_total_capital"
]

ES_FIELDS = [
    "snapshot_date", "shareholder_name", "shares_held",
    "capital_contribution", "shareholding_ratio", "snapshot_type"
]


def normalize_name(name):
    """Normalize company/subsidiary names for matching"""
    name = name.strip()
    name = re.sub(r'[（(].*?[）)]', '', name)  # remove parenthetical
    return name


def match_records(auto_rows, gold_rows, record_type):
    """Align auto and gold records by date + name"""
    comparisons = []
    matched_gold = set()
    matched_auto = set()

    # Index gold by (date, name)
    gold_idx = {}
    for i, g in enumerate(gold_rows):
        if record_type == "subscription_flow":
            key = (g.get("subscription_date", ""), normalize_name(g.get("subscriber_name", "")))
        else:
            key = (g.get("snapshot_date", ""), normalize_name(g.get("shareholder_name", "")))
        gold_idx.setdefault(key, []).append((i, g))

    # Match auto records against gold
    for j, a in enumerate(auto_rows):
        if record_type == "subscription_flow":
            key = (a.get("subscription_date", ""), normalize_name(a.get("subscriber_name", "")))
        else:
            key = (a.get("snapshot_date", ""), normalize_name(a.get("shareholder_name", "")))

        if key in gold_idx and gold_idx[key]:
            gi, g = gold_idx[key].pop(0)
            matched_gold.add(gi)
            matched_auto.add(j)
            comparisons.append(("matched", a, g))
        else:
            comparisons.append(("auto_only", a, None))

    # Unmatched gold records
    for i, g in enumerate(gold_rows):
        if i not in matched_gold:
            comparisons.append(("gold_only", None, g))

    return comparisons


def compare_fields(auto_rec, gold_rec, fields):
    """Compare field values between auto and gold records"""
    results = []
    for fld in fields:
        av = auto_rec.get(fld) if auto_rec else None
        gv = gold_rec.get(fld) if gold_rec else None

        # Normalize None vs empty string
        if av == "":
            av = None
        if gv == "":
            gv = None

        if av is None and gv is None:
            status = "both_null"
        elif av is None and gv is not None:
            status = "auto_missing"
        elif av is not None and gv is None:
            status = "gold_missing"
        elif isinstance(av, (int, float)) and isinstance(gv, (int, float)):
            # Float comparison with tolerance
            if abs(av - gv) < 0.01:
                status = "match"
            else:
                status = "mismatch"
        elif str(av) == str(gv):
            status = "match"
        else:
            status = "mismatch"

        results.append({
            "field": fld,
            "auto_value": str(av)[:100] if av is not None else "",
            "gold_value": str(gv)[:100] if gv is not None else "",
            "status": status
        })
    return results


def main():
    print("=" * 60)
    print("Auto vs Gold 对比")
    print("=" * 60)

    # Load gold JSONL
    gold_data = {}
    for f in sorted(GOLD_DIR.glob("*.jsonl")):
        with open(f, encoding="utf-8") as fh:
            rows = [json.loads(ln) for ln in fh if ln.strip()]
        company = f.stem
        sf = [r for r in rows if r.get("record_type") == "subscription_flow"]
        es = [r for r in rows if r.get("record_type") == "equity_snapshot"]
        gold_data[company] = {"subscription_flow": sf, "equity_snapshot": es}
        print(f"  Gold: {company}: {len(sf)} flows + {len(es)} snaps")

    # Run auto extraction
    auto_data = {}
    if HAS_AUTO:
        print("\n>>> 自动提取中...")
        auto_results = auto_extract_all()
        for company_key, result in auto_results.items():
            company_match = [(c, k) for (f, c, k), files in COMPANY_MD_MAP.items() if k == company_key]
            stock_code = company_match[0][0] if company_match else "unknown"
            auto_data[stock_code] = {"subscription_flow": result.get("flows", []),
                                     "equity_snapshot": result.get("snaps", [])}
            print(f"  Auto: {company_key}({stock_code}): {len(result.get('flows',[]))} flows + {len(result.get('snaps',[]))} snaps")
    else:
        print("\n  自动提取不可用 (auto_extract import failed)")

    # Compare
    all_sf_results = []
    all_es_results = []
    stats = {}

    for company, gd in gold_data.items():
        ad = auto_data.get(company, {"subscription_flow": [], "equity_snapshot": []})
        code = company.split("_")[0] if "_" in company else company

        # subscription_flow comparison
        sf_comps = match_records(ad["subscription_flow"], gd["subscription_flow"], "subscription_flow")
        sf_matches = sum(1 for t, _, _ in sf_comps if t == "matched")
        sf_auto_only = sum(1 for t, _, _ in sf_comps if t == "auto_only")
        sf_gold_only = sum(1 for t, _, _ in sf_comps if t == "gold_only")

        for match_type, auto_rec, gold_rec in sf_comps:
            if match_type == "matched":
                field_results = compare_fields(auto_rec, gold_rec, SF_FIELDS)
                for fr in field_results:
                    all_sf_results.append({
                        "company": company, "match_type": match_type,
                        "date": gold_rec.get("subscription_date", ""),
                        "name": gold_rec.get("subscriber_name", ""),
                        **fr
                    })
            elif match_type == "auto_only":
                for fld in SF_FIELDS:
                    all_sf_results.append({
                        "company": company, "match_type": "auto_only(误提)",
                        "date": auto_rec.get("subscription_date", ""),
                        "name": auto_rec.get("subscriber_name", ""),
                        "field": fld, "auto_value": str(auto_rec.get(fld, ""))[:100],
                        "gold_value": "(无)", "status": "auto_only"
                    })
            else:
                for fld in SF_FIELDS:
                    all_sf_results.append({
                        "company": company, "match_type": "gold_only(漏提)",
                        "date": gold_rec.get("subscription_date", ""),
                        "name": gold_rec.get("subscriber_name", ""),
                        "field": fld, "auto_value": "(无)",
                        "gold_value": str(gold_rec.get(fld, ""))[:100], "status": "gold_only"
                    })

        # equity_snapshot comparison
        es_comps = match_records(ad["equity_snapshot"], gd["equity_snapshot"], "equity_snapshot")
        es_matches = sum(1 for t, _, _ in es_comps if t == "matched")
        es_auto_only = sum(1 for t, _, _ in es_comps if t == "auto_only")
        es_gold_only = sum(1 for t, _, _ in es_comps if t == "gold_only")

        for match_type, auto_rec, gold_rec in es_comps:
            if match_type == "matched":
                field_results = compare_fields(auto_rec, gold_rec, ES_FIELDS)
                for fr in field_results:
                    all_es_results.append({
                        "company": company, "match_type": match_type,
                        "date": gold_rec.get("snapshot_date", ""),
                        "name": gold_rec.get("shareholder_name", ""),
                        **fr
                    })
            elif match_type == "auto_only":
                for fld in ES_FIELDS:
                    all_es_results.append({
                        "company": company, "match_type": "auto_only(误提)",
                        "date": auto_rec.get("snapshot_date", ""),
                        "name": auto_rec.get("shareholder_name", ""),
                        "field": fld, "auto_value": str(auto_rec.get(fld, ""))[:100],
                        "gold_value": "(无)", "status": "auto_only"
                    })
            else:
                for fld in ES_FIELDS:
                    all_es_results.append({
                        "company": company, "match_type": "gold_only(漏提)",
                        "date": gold_rec.get("snapshot_date", ""),
                        "name": gold_rec.get("shareholder_name", ""),
                        "field": fld, "auto_value": "(无)",
                        "gold_value": str(gold_rec.get(fld, ""))[:100], "status": "gold_only"
                    })

        total_sf = len(gd["subscription_flow"])
        total_es = len(gd["equity_snapshot"])
        stats[company] = {
            "code": code,
            "sf_total": total_sf, "sf_auto_matched": sf_matches,
            "sf_auto_only": sf_auto_only, "sf_gold_only": sf_gold_only,
            "es_total": total_es, "es_auto_matched": es_matches,
            "es_auto_only": es_auto_only, "es_gold_only": es_gold_only,
        }

        print(f"\n  {company}:")
        print(f"    subscription_flow: gold={total_sf}, auto_matched={sf_matches}, auto_only={sf_auto_only}, gold_only={sf_gold_only}")
        print(f"    equity_snapshot:   gold={total_es}, auto_matched={es_matches}, auto_only={es_auto_only}, gold_only={es_gold_only}")

    # Statistics summary
    sf_field_stats = defaultdict(lambda: {"match": 0, "mismatch": 0, "auto_missing": 0, "gold_missing": 0, "both_null": 0})
    for r in all_sf_results:
        if r["match_type"] == "matched":
            sf_field_stats[r["field"]][r["status"]] += 1

    es_field_stats = defaultdict(lambda: {"match": 0, "mismatch": 0, "auto_missing": 0, "gold_missing": 0, "both_null": 0})
    for r in all_es_results:
        if r["match_type"] == "matched":
            es_field_stats[r["field"]][r["status"]] += 1

    # Export Excel
    if HAS_XL:
        wb = openpyxl.Workbook()
        hf = Font(bold=True, size=11)
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        hfont = Font(bold=True, size=11, color="FFFFFF")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Sheet 1: subscription_flow
        ws1 = wb.active
        ws1.title = "subscription_flow对比"
        sf_headers = ["公司", "匹配类型", "日期", "认购方/股东", "字段", "自动值", "Gold值", "状态"]
        for c, h in enumerate(sf_headers, 1):
            cell = ws1.cell(row=1, column=c, value=h)
            cell.font = hfont; cell.fill = hfill
        for i, r in enumerate(all_sf_results, 2):
            ws1.cell(row=i, column=1, value=r["company"])
            ws1.cell(row=i, column=2, value=r["match_type"])
            ws1.cell(row=i, column=3, value=r["date"])
            ws1.cell(row=i, column=4, value=r["name"])
            ws1.cell(row=i, column=5, value=r["field"])
            ws1.cell(row=i, column=6, value=r["auto_value"])
            ws1.cell(row=i, column=7, value=r["gold_value"])
            status_cell = ws1.cell(row=i, column=8, value=r["status"])
            if r["status"] == "mismatch":
                status_cell.fill = red_fill
            elif r["status"] == "match":
                status_cell.fill = green_fill
            elif "missing" in r["status"] or "only" in r["status"]:
                status_cell.fill = yellow_fill

        # Sheet 2: equity_snapshot
        ws2 = wb.create_sheet("equity_snapshot对比")
        for c, h in enumerate(sf_headers, 1):
            cell = ws2.cell(row=1, column=c, value=h)
            cell.font = hfont; cell.fill = hfill
        for i, r in enumerate(all_es_results, 2):
            ws2.cell(row=i, column=1, value=r["company"])
            ws2.cell(row=i, column=2, value=r["match_type"])
            ws2.cell(row=i, column=3, value=r["date"])
            ws2.cell(row=i, column=4, value=r["name"])
            ws2.cell(row=i, column=5, value=r["field"])
            ws2.cell(row=i, column=6, value=r["auto_value"])
            ws2.cell(row=i, column=7, value=r["gold_value"])
            status_cell = ws2.cell(row=i, column=8, value=r["status"])
            if r["status"] == "mismatch":
                status_cell.fill = red_fill
            elif r["status"] == "match":
                status_cell.fill = green_fill
            elif "missing" in r["status"] or "only" in r["status"]:
                status_cell.fill = yellow_fill

        # Sheet 3: Statistics
        ws3 = wb.create_sheet("统计")
        ws3.cell(row=1, column=1, value="公司").font = hfont
        ws3.cell(row=1, column=1).fill = hfill
        for c, h in enumerate(["代码", "SF_Gold总数", "SF_Auto匹配", "SF_Auto-only", "SF_Gold-only",
                                "ES_Gold总数", "ES_Auto匹配", "ES_Auto-only", "ES_Gold-only"], 2):
            cell = ws3.cell(row=1, column=c, value=h)
            cell.font = hfont; cell.fill = hfill

        for i, (company, st) in enumerate(sorted(stats.items()), 2):
            ws3.cell(row=i, column=1, value=company)
            ws3.cell(row=i, column=2, value=st["code"])
            ws3.cell(row=i, column=3, value=st["sf_total"])
            ws3.cell(row=i, column=4, value=st["sf_auto_matched"])
            ws3.cell(row=i, column=5, value=st["sf_auto_only"])
            ws3.cell(row=i, column=6, value=st["sf_gold_only"])
            ws3.cell(row=i, column=7, value=st["es_total"])
            ws3.cell(row=i, column=8, value=st["es_auto_matched"])
            ws3.cell(row=i, column=9, value=st["es_auto_only"])
            ws3.cell(row=i, column=10, value=st["es_gold_only"])

        # Field-level stats
        row = len(stats) + 4
        ws3.cell(row=row, column=1, value="字段级准确率 (subscription_flow)").font = Font(bold=True, size=12)
        row += 1
        for c, h in enumerate(["字段", "匹配", "不匹配", "Auto缺失", "Gold缺失", "双Null", "准确率"], 1):
            cell = ws3.cell(row=row, column=c, value=h)
            cell.font = hfont; cell.fill = hfill
        row += 1
        for fld in SF_FIELDS:
            st = sf_field_stats[fld]
            total = sum(st.values())
            acc = st["match"] / max(total - st["both_null"], 1) if total > 0 else 0
            ws3.cell(row=row, column=1, value=fld)
            ws3.cell(row=row, column=2, value=st["match"])
            ws3.cell(row=row, column=3, value=st["mismatch"])
            ws3.cell(row=row, column=4, value=st["auto_missing"])
            ws3.cell(row=row, column=5, value=st["gold_missing"])
            ws3.cell(row=row, column=6, value=st["both_null"])
            ws3.cell(row=row, column=7, value=f"{acc:.1%}")
            row += 1

        row += 1
        ws3.cell(row=row, column=1, value="字段级准确率 (equity_snapshot)").font = Font(bold=True, size=12)
        row += 1
        for c, h in enumerate(["字段", "匹配", "不匹配", "Auto缺失", "Gold缺失", "双Null", "准确率"], 1):
            cell = ws3.cell(row=row, column=c, value=h)
            cell.font = hfont; cell.fill = hfill
        row += 1
        for fld in ES_FIELDS:
            st = es_field_stats[fld]
            total = sum(st.values())
            acc = st["match"] / max(total - st["both_null"], 1) if total > 0 else 0
            ws3.cell(row=row, column=1, value=fld)
            ws3.cell(row=row, column=2, value=st["match"])
            ws3.cell(row=row, column=3, value=st["mismatch"])
            ws3.cell(row=row, column=4, value=st["auto_missing"])
            ws3.cell(row=row, column=5, value=st["gold_missing"])
            ws3.cell(row=row, column=6, value=st["both_null"])
            ws3.cell(row=row, column=7, value=f"{acc:.1%}")
            row += 1

        out_path = REPORTS_DIR / "auto_vs_gold_comparison.xlsx"
        wb.save(out_path)
        print(f"\n✓ Excel: {out_path}")

    # Summary JSON
    summary = {
        "generated_at": datetime.now().isoformat(),
        "companies": stats,
        "sf_field_accuracy": {
            fld: {
                "match": st["match"], "mismatch": st["mismatch"],
                "auto_missing": st["auto_missing"], "gold_missing": st["gold_missing"],
                "both_null": st["both_null"]
            } for fld, st in sf_field_stats.items()
        },
        "es_field_accuracy": {
            fld: {
                "match": st["match"], "mismatch": st["mismatch"],
                "auto_missing": st["auto_missing"], "gold_missing": st["gold_missing"],
                "both_null": st["both_null"]
            } for fld, st in es_field_stats.items()
        }
    }
    summary_path = REPORTS_DIR / "auto_vs_gold_summary.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"✓ Summary: {summary_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
