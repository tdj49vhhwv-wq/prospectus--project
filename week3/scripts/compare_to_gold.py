#!/usr/bin/env python3
"""
Auto vs Gold 逐字段对比

输入:
  - week3/outputs/auto_jsonl/ (自动提取结果)
  - week3/manual_gold/ (人工Gold)
输出:
  - week3/evaluation/auto_vs_gold_comparison.xlsx
  - week3/evaluation/auto_vs_gold_summary.json
"""
import sys, json, re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent))
from pipeline_config import *

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_XL = True
except ImportError:
    HAS_XL = False

SF_FIELDS = ["subscription_date", "subscriber_name", "shares_subscribed",
             "amount_subscribed", "price_per_share", "event_context"]
ST_FIELDS = ["transfer_date", "transferor_name", "transferee_name",
             "shares_transferred", "transfer_amount", "transfer_type"]
ES_FIELDS = ["snapshot_date", "shareholder_name", "shares_held",
             "capital_contribution", "shareholding_ratio", "snapshot_type"]


def normalize(s):
    return re.sub(r'[（(].*?[）)]', '', (s or '').strip())


def load_jsonl(path):
    if not path.exists(): return []
    with open(path, encoding="utf-8") as f:
        return [json.loads(ln) for ln in f if ln.strip()]


def match_and_compare(auto_rows, gold_rows, record_type):
    """对齐auto和gold记录,逐字段对比"""
    comparisons = []
    gold_idx = {}
    fields = SF_FIELDS if record_type == "subscription_flow" else (
        ST_FIELDS if record_type == "share_transfer_flow" else ES_FIELDS)

    date_key = "subscription_date" if record_type == "subscription_flow" else (
        "transfer_date" if record_type == "share_transfer_flow" else "snapshot_date")
    name_key = "subscriber_name" if record_type == "subscription_flow" else (
        "transferee_name" if record_type == "share_transfer_flow" else "shareholder_name")

    for i, g in enumerate(gold_rows):
        key = (g.get(date_key, ""), normalize(g.get(name_key, "")))
        gold_idx.setdefault(key, []).append((i, g))

    matched_gold = set()
    for a in auto_rows:
        key = (a.get(date_key, ""), normalize(a.get(name_key, "")))
        if key in gold_idx and gold_idx[key]:
            gi, g = gold_idx[key].pop(0)
            matched_gold.add(gi)
            for fld in fields:
                av, gv = a.get(fld), g.get(fld)
                if av is None and gv is None: st = "both_null"
                elif av is None: st = "auto_missing"
                elif gv is None: st = "gold_missing"
                elif isinstance(av, (int, float)) and isinstance(gv, (int, float)):
                    st = "match" if abs(av - gv) < 0.01 else "mismatch"
                else:
                    st = "match" if str(av) == str(gv) else "mismatch"
                comparisons.append({"company": "", "match_type": "matched",
                    "date": g.get(date_key, ""), "name": g.get(name_key, ""),
                    "field": fld, "auto_value": str(av)[:100] if av is not None else "",
                    "gold_value": str(gv)[:100] if gv is not None else "", "status": st})
        else:
            for fld in fields:
                comparisons.append({"company": "", "match_type": "auto_only(误提)",
                    "date": a.get(date_key, ""), "name": a.get(name_key, ""),
                    "field": fld, "auto_value": str(a.get(fld, ""))[:100],
                    "gold_value": "(无)", "status": "auto_only"})

    for i, g in enumerate(gold_rows):
        if i not in matched_gold:
            for fld in fields:
                comparisons.append({"company": "", "match_type": "gold_only(漏提)",
                    "date": g.get(date_key, ""), "name": g.get(name_key, ""),
                    "field": fld, "auto_value": "(无)",
                    "gold_value": str(g.get(fld, ""))[:100], "status": "gold_only"})

    return comparisons


def main():
    print("=" * 60)
    print("[AUTO] compare_to_gold — Auto vs Gold对比")
    print("=" * 60)

    # 加载Gold
    gold_sf = load_jsonl(SF_GOLD)
    gold_st = load_jsonl(ST_GOLD)
    gold_es = load_jsonl(ES_GOLD)
    print(f"Gold: {len(gold_sf)} sf + {len(gold_st)} st + {len(gold_es)} es")

    # 加载Auto
    auto_sf = load_jsonl(AUTO_SF_JSONL)
    auto_st = load_jsonl(AUTO_ST_JSONL)
    auto_es = load_jsonl(AUTO_ES_JSONL)
    print(f"Auto: {len(auto_sf)} sf + {len(auto_st)} st + {len(auto_es)} es")

    all_results = []
    for label, auto, gold, rt in [
        ("subscription_flow", auto_sf, gold_sf, "subscription_flow"),
        ("share_transfer_flow", auto_st, gold_st, "share_transfer_flow"),
        ("equity_snapshot", auto_es, gold_es, "equity_snapshot"),
    ]:
        comps = match_and_compare(auto, gold, rt)
        matched = sum(1 for c in comps if c["match_type"] == "matched" and c["status"] == "match")
        mismatched = sum(1 for c in comps if c["status"] == "mismatch")
        auto_only = sum(1 for c in comps if c["match_type"] == "auto_only(误提)")
        gold_only = sum(1 for c in comps if c["match_type"] == "gold_only(漏提)")
        print(f"  {label}: matched={matched} mismatched={mismatched} auto_only={auto_only} gold_only={gold_only}")
        for c in comps:
            c["company"] = label
        all_results.extend(comps)

    # 输出Excel
    if HAS_XL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "field_comparison"
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        headers = ["类型", "匹配类型", "日期", "名称", "字段", "自动值", "Gold值", "状态"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hf; cell.fill = hfill
        for i, r in enumerate(all_results, 2):
            ws.cell(row=i, column=1, value=r["company"])
            ws.cell(row=i, column=2, value=r["match_type"])
            ws.cell(row=i, column=3, value=r["date"])
            ws.cell(row=i, column=4, value=r["name"])
            ws.cell(row=i, column=5, value=r["field"])
            ws.cell(row=i, column=6, value=r["auto_value"])
            ws.cell(row=i, column=7, value=r["gold_value"])
            ws.cell(row=i, column=8, value=r["status"])
        wb.save(COMPARISON_XLSX)
        print(f"✓ {COMPARISON_XLSX.name}")

    # 输出Summary JSON
    summary = {
        "generated_at": datetime.now().isoformat(),
        "gold_counts": {"subscription_flow": len(gold_sf), "share_transfer_flow": len(gold_st), "equity_snapshot": len(gold_es)},
        "auto_counts": {"subscription_flow": len(auto_sf), "share_transfer_flow": len(auto_st), "equity_snapshot": len(auto_es)},
        "total_comparisons": len(all_results),
    }
    with open(COMPARISON_JSON, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"✓ {COMPARISON_JSON.name}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
